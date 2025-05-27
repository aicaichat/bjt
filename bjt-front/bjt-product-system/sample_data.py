#!/usr/bin/env python3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def create_sample_excel():
    """Create a sample Excel file for testing Excel-to-SQL conversion"""
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Add instructions sheet
    ws = wb.create_sheet("使用说明")
    ws['A1'] = "Sample Excel for Testing"
    ws['A1'].font = Font(bold=True, size=16)
    
    # Add a sample product_lines sheet
    ws = wb.create_sheet("wp_bjt_product_lines")
    
    # Add column headers
    headers = ['id', 'title_zh', 'title_en', 'description_zh', 'description_en', 'image_url', 'code', 'status', 'sort_order']
    for i, header in enumerate(headers):
        cell = ws.cell(row=1, column=i+1, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Add sample data
    data = [
        [5, '新产品线', 'New Product Line', '这是一个新产品线描述', 'This is a new product line description', '/images/shop/new.jpg', 'new_line', 'publish', 50],
        [6, '测试产品线', 'Test Product Line', '测试产品线描述', 'Test product line description', '/images/shop/test.jpg', 'test_line', 'publish', 60],
        [7, 'NULL', 'NULL', 'NULL', 'NULL', 'NULL', 'NULL', 'draft', 70]
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value if value != 'NULL' else None)
    
    # Add a sample shapes sheet
    ws = wb.create_sheet("wp_bjt_shapes")
    
    # Add column headers
    headers = ['id', 'product_line_id', 'code', 'name_en', 'name_zh', 'image_url', 'status', 'sort_order']
    for i, header in enumerate(headers):
        cell = ws.cell(row=1, column=i+1, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # Add sample data
    data = [
        [4, 1, 'square', 'square', '方形', '/images/shop/square.jpg', 'publish', 40],
        [5, 2, 'triangle', 'triangle', '三角形', '/images/shop/triangle.jpg', 'publish', 50],
        [6, 3, 'hexagon', 'hexagon', '六边形', '/images/shop/hexagon.jpg', 'draft', 60]
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save the workbook
    output_path = 'sample_data.xlsx'
    wb.save(output_path)
    print(f"Sample Excel file created at: {output_path}")
    return output_path

if __name__ == "__main__":
    create_sample_excel() 