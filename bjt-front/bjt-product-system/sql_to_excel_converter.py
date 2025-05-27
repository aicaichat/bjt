import re
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import json

# 字段默认值映射
FIELD_DEFAULTS = {
    'created_at': 'NOW()',
    'updated_at': 'NOW()',
    'status': "'publish'",
    'sort_order': 0,
    'image_url': "''",
    'description_zh': "''",
    'description_en': "''",
    'type': "''",
    'code': "''",
    'brand': "''",
    'spec': "''",
    'spec_imperial': "''",
    'unit': "'pcs'",
    # 可根据实际表结构继续补充
}

class SQLExcelConverter:
    def __init__(self, sql_dir=None, output_dir=None):
        self.sql_dir = sql_dir or os.path.join(os.getcwd(), 'docker', 'dev', 'mysql')
        self.output_dir = output_dir or os.getcwd()
        self.tables = {}
        self.warnings = []
        self.table_schemas = {}  # 存储表结构定义
        
    def parse_table_schemas(self):
        """从init.sql解析表结构定义"""
        init_sql_path = os.path.join(self.sql_dir, 'init.sql')
        if not os.path.exists(init_sql_path):
            print(f"Warning: init.sql not found at {init_sql_path}. Will use INSERT statements for table structure.")
            return
            
        try:
            with open(init_sql_path, 'r', encoding='utf-8') as f:
                sql_content = f.read()
                
            # 解析CREATE TABLE语句
            create_pattern = r'CREATE\s+TABLE\s+IF\s+NOT\s+EXISTS\s+`(\w+)`\s*\(([\s\S]*?)\)\s*ENGINE'
            table_matches = re.finditer(create_pattern, sql_content, re.IGNORECASE)
            
            for match in table_matches:
                table_name = match.group(1)
                columns_def = match.group(2)
                
                # 从列定义中提取字段名
                column_pattern = r'`(\w+)`\s+\w+'
                columns = re.findall(column_pattern, columns_def)
                
                self.table_schemas[table_name] = columns
                print(f"Parsed schema for table {table_name}: {len(columns)} columns")
                
        except Exception as e:
            print(f"Error parsing init.sql: {str(e)}")
            
    def parse_sql_files(self, file_paths=None):
        """Parse SQL files and extract table structures and data"""
        # 先解析表结构定义
        self.parse_table_schemas()
        
        if file_paths is None:
            # Use all SQL files in the directory
            file_paths = [os.path.join(self.sql_dir, f) for f in os.listdir(self.sql_dir) 
                         if f.endswith('.sql') and f != 'init.sql']
        
        for file_path in file_paths:
            self._parse_sql_file(file_path)
            
        # 确保所有表都有完整的字段列表
        self._ensure_complete_columns()
            
        print(f"Successfully parsed {len(self.tables)} tables from SQL files.")
        return self.tables
    
    def _ensure_complete_columns(self):
        """确保所有料号表都有完整的字段列表"""
        critical_tables = ['wp_bjt_parts', 'wp_bjt_accessories', 'wp_bjt_consumables', 'wp_bjt_spare_parts']
        
        for table_name in critical_tables:
            if table_name in self.tables and table_name in self.table_schemas:
                # 获取完整字段列表
                complete_columns = self.table_schemas[table_name]
                # 获取当前字段列表
                current_columns = self.tables[table_name]['columns']
                
                # 检查是否缺少字段
                missing_columns = [col for col in complete_columns if col not in current_columns]
                
                if missing_columns:
                    print(f"Adding {len(missing_columns)} missing columns to {table_name}")
                    
                    # 更新字段列表
                    self.tables[table_name]['columns'] = complete_columns
                    
                    # 更新数据行，为缺失字段补NULL
                    for i, row in enumerate(self.tables[table_name]['rows']):
                        # 构建完整行数据
                        complete_row = []
                        for col in complete_columns:
                            if col in current_columns:
                                # 保留原有字段的值
                                idx = current_columns.index(col)
                                if idx < len(row):
                                    complete_row.append(row[idx])
                                else:
                                    complete_row.append(None)
                            else:
                                # 补充缺失字段
                                complete_row.append(None)
                        
                        # 更新行数据
                        self.tables[table_name]['rows'][i] = complete_row
                        
    def _parse_sql_file(self, file_path):
        """Parse a single SQL file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                sql_content = f.read()
                
            # Extract INSERT statements
            insert_pattern = r'INSERT\s+INTO\s+`?(\w+)`?\s*\((.*?)\)\s*VALUES\s*([\s\S]*?)(?=INSERT|$|;)'
            matches = re.finditer(insert_pattern, sql_content, re.IGNORECASE)
            
            for match in matches:
                table_name = match.group(1)
                columns_str = match.group(2)
                values_str = match.group(3)
                
                # Parse columns
                columns = [col.strip().replace('`', '') for col in columns_str.split(',')]
                
                # Parse values
                values_pattern = r'\((.*?)\)'
                values_matches = re.finditer(values_pattern, values_str)
                rows = []
                
                for val_match in values_matches:
                    val_str = val_match.group(1)
                    # Split by comma but respect quoted strings
                    values = []
                    current_value = ""
                    in_quote = False
                    quote_char = None
                    for char in val_str + ',':  # Add comma to process the last value
                        if char in ["'", '"'] and (not in_quote or char == quote_char):
                            in_quote = not in_quote
                            if in_quote:
                                quote_char = char
                            else:
                                quote_char = None
                            current_value += char
                        elif char == ',' and not in_quote:
                            values.append(current_value.strip())
                            current_value = ""
                        else:
                            current_value += char
                    # Clean up values
                    cleaned_values = []
                    for val in values:
                        val = val.strip()
                        if val.upper() == 'NULL':
                            cleaned_values.append(None)
                        elif val.upper() == 'NOW()':
                            cleaned_values.append('NOW()')
                        elif (val.startswith("'") and val.endswith("'")) or (val.startswith('"') and val.endswith('"')):
                            cleaned_values.append(val[1:-1])
                        else:
                            escaped_val = str(val).replace("'", "''")
                            # Fix image paths: replace backslashes with forward slashes
                            if ('image' in col.lower() or 'url' in col.lower()) and '\\' in escaped_val:
                                escaped_val = escaped_val.replace('\\', '/')
                            cleaned_values.append(f"'{escaped_val}'")
                    # 修复字段数量不匹配：补齐或截断
                    if len(cleaned_values) < len(columns):
                        # 补齐缺失字段
                        cleaned_values += [None] * (len(columns) - len(cleaned_values))
                    elif len(cleaned_values) > len(columns):
                        # 截断多余字段
                        cleaned_values = cleaned_values[:len(columns)]
                    # 现在字段数量严格等于表结构
                    rows.append(cleaned_values)
                # Add or update table data
                if table_name not in self.tables:
                    self.tables[table_name] = {
                        'columns': columns,
                        'rows': rows
                    }
                else:
                    self.tables[table_name]['rows'].extend(rows)
        except Exception as e:
            print(f"Error parsing SQL file {file_path}: {str(e)}")
            raise
    
    def generate_excel_template(self, output_path=None):
        """Generate Excel template with sheets for each table"""
        if not self.tables:
            raise ValueError("No tables parsed. Please parse SQL files first.")
        
        if output_path is None:
            output_path = os.path.join(self.output_dir, 'database_import_template.xlsx')
        
        # Create a new workbook
        wb = Workbook()
        
        # Remove the default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Add instructions sheet
        self._add_instructions_sheet(wb)
        
        # Add sheets for each table
        for table_name, table_data in self.tables.items():
            self._add_table_sheet(wb, table_name, table_data)
        
        # Save the workbook
        wb.save(output_path)
        print(f"Excel template generated at: {output_path}")
        return output_path
    
    def _add_instructions_sheet(self, workbook):
        """Add instructions sheet to the workbook"""
        ws = workbook.create_sheet("使用说明")
        
        # Set column widths
        ws.column_dimensions['A'].width = 100
        
        # Title
        ws['A1'] = "数据库导入模板使用说明"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Instructions
        instructions = [
            "",
            "1. 本Excel文件包含多个工作表，每个工作表对应一个数据库表。",
            "2. 每个工作表的第一行是列名，对应数据库表的字段名。",
            "3. 第2-6行是留给用户填写数据的空行。可自行扩展到多行",
            "4. 第8行开始是示例数据，仅供参考，不会被导入。",
            "5. 填写数据时，请遵循以下规则：",
            "   - 数字类型直接填写数字，不需要引号",
            "   - 字符串类型直接填写文本，不需要引号",
            "   - 日期类型请使用YYYY-MM-DD格式",
            "   - 日期时间类型请使用YYYY-MM-DD HH:MM:SS格式",
            "   - 空值请留空或填写NULL",
            "   - 如需使用数据库函数如NOW()，请直接填写函数名",
            "",
            "6. 填写完毕后，使用导入脚本将数据导入数据库。",
            "",
            "注意：请不要修改表格的结构，包括列名和格式。",
            "",
            "导出 SQL 前请删除或清空所有样例数据行（通常为第8行及以后），否则这些数据也会被转为 SQL 语句！",
            "",
            "图片字段填写要求：",
            "- 图片字段请填写图片的URL路径，如 /images/shop/LA-E4S.jpg，不要直接粘贴图片。",
            "- 推荐图片尺寸：宽度800像素，高度800像素，比例要求1:1（正方形）。",
            "- 单张图片大小建议不超过500KB，最佳为100~300KB，以保证前端页面加载速度。",
            "- 推荐图片格式：JPG、PNG（如需透明背景请用PNG）。",
            "- 图片文件名建议使用英文、数字、下划线，避免中文和特殊字符。",
            "- 图片应居中、主体清晰，背景简洁，避免过多留白或裁切不当。",
            "- 如有多张图片字段（如image1_url, image2_url），请分别填写对应图片的URL。",
        ]
        
        for i, text in enumerate(instructions):
            ws[f'A{i+2}'] = text
            if text.startswith("   -"):
                ws[f'A{i+2}'].font = Font(italic=True)
            elif not text:
                continue
            elif text[0].isdigit():
                ws[f'A{i+2}'].font = Font(bold=True)
        
        return ws
    
    def _add_table_sheet(self, workbook, table_name, table_data):
        """Add a sheet for the table to the workbook"""
        ws = workbook.create_sheet(table_name)
        
        columns = table_data['columns']
        rows = table_data['rows']
        
        # Set column widths
        for i, col in enumerate(columns):
            col_letter = get_column_letter(i + 1)
            ws.column_dimensions[col_letter].width = max(15, min(30, len(col) + 5))
        
        # Header row
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        header_font = Font(bold=True)
        header_border = Border(
            bottom=Side(style='medium'),
            top=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        
        for i, col in enumerate(columns):
            cell = ws.cell(row=1, column=i+1, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Empty rows for data entry
        data_border = Border(
            bottom=Side(style='thin'),
            top=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        
        for row_idx in range(2, 7):
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx, value="")
                cell.border = data_border
        
        # Separator row
        ws.cell(row=7, column=1, value="=== 示例数据 ===")
        ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=len(columns))
        ws.cell(row=7, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=7, column=1).font = Font(bold=True, italic=True)
        ws.cell(row=7, column=1).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Example data rows
        example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        example_border = Border(
            bottom=Side(style='thin'),
            top=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        
        for row_idx, row_data in enumerate(rows[:5], start=8):
            for col_idx, (col, value) in enumerate(zip(columns, row_data), start=1):
                # 自动填充默认值
                if value is None or value == '':
                    value = FIELD_DEFAULTS.get(col, value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = example_fill
                cell.border = example_border
        
        return ws
    
    def generate_import_script(self, excel_path=None, output_path=None):
        """Generate Python script to import data from Excel to database"""
        if excel_path is None:
            excel_path = os.path.join(self.output_dir, 'database_import_template.xlsx')
        
        if output_path is None:
            output_path = os.path.join(self.output_dir, 'import_excel.py')
        
        script_content = """#!/usr/bin/env python3
import os
import pandas as pd
import mysql.connector
import getpass
import sys

def main():
    # Get database connection info
    print("=== Database Import Tool ===")
    host = input("Database host [localhost]: ") or "localhost"
    port = input("Database port [3306]: ") or "3306"
    user = input("Database user [root]: ") or "root"
    password = getpass.getpass("Database password: ")
    database = input("Database name: ")
    
    if not database:
        print("Error: Database name is required.")
        sys.exit(1)
    
    # Connect to database
    try:
        conn = mysql.connector.connect(
            host=host,
            port=port,
            user=user,
            password=password,
            database=database
        )
        cursor = conn.cursor()
        print(f"Connected to database {database} on {host}:{port}")
    except Exception as e:
        print(f"Error connecting to database: {str(e)}")
        sys.exit(1)
    
    # Get Excel file path
    excel_path = input(f"Excel file path [{os.path.basename(r'__EXCEL_PATH__')}]: ") or r'__EXCEL_PATH__'
    
    try:
        # Read Excel file
        xl = pd.ExcelFile(excel_path)
        
        # Process each sheet (table)
        for sheet_name in xl.sheet_names:
            if sheet_name == '使用说明':
                continue
                
            print(f"\nProcessing table: {sheet_name}")
            
            # Read sheet data
            df = pd.read_excel(excel_path, sheet_name=sheet_name)
            
            # Get only the first 5 rows (user data)
            df = df.iloc[:5]
            
            # Remove empty rows
            df = df.dropna(how='all')
            
            if df.empty:
                print(f"No data to import for table {sheet_name}")
                continue
            
            # Process each row
            for _, row in df.iterrows():
                # Replace NaN with None
                values = [None if pd.isna(val) else val for val in row]
                
                # Generate placeholders
                placeholders = ", ".join(["%s"] * len(values))
                
                # Generate column names
                columns = ", ".join([f"`{col}`" for col in df.columns])
                
                # Generate SQL
                sql = f"INSERT INTO `{sheet_name}` ({columns}) VALUES ({placeholders})"
                
                try:
                    cursor.execute(sql, values)
                    print(f"Inserted row into {sheet_name}")
                except Exception as e:
                    print(f"Error inserting row into {sheet_name}: {str(e)}")
        
        # Commit changes
        conn.commit()
        print("\nAll data imported successfully!")
        
    except Exception as e:
        print(f"Error processing Excel file: {str(e)}")
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()
            print("Database connection closed.")

if __name__ == "__main__":
    main()
"""
        
        # Replace placeholder with actual Excel path
        script_content = script_content.replace('__EXCEL_PATH__', excel_path)
        
        # Write script to file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(script_content)
        
        # Make script executable
        os.chmod(output_path, 0o755)
        
        print(f"Import script generated at: {output_path}")
        return output_path

    def excel_to_sql(self, excel_path, output_path=None, dialect='mysql', insert_format='multi-row', batch_size=1000, on_duplicate=False):
        """Convert Excel file to SQL INSERT statements"""
        if output_path is None:
            # Generate a default output path based on the Excel filename
            excel_filename = os.path.basename(excel_path)
            # Sanitize filename: remove "database_import_templatebak", "（", "）", ".xlsx" and replace with underscores
            sanitized_name = excel_filename.replace('database_import_templatebak', '')
            sanitized_name = sanitized_name.replace('（', '_').replace('）', '').replace(' ', '_')
            sanitized_name = os.path.splitext(sanitized_name)[0] # remove .xlsx
            if not sanitized_name or sanitized_name == '_': # Handle cases where the name becomes empty or just an underscore
                 sanitized_name = "excel_import" # Default if sanitization results in empty string
            
            # Ensure the output directory exists
            generated_sql_dir = os.path.join(self.output_dir or os.getcwd(), 'generated_sql_imports')
            os.makedirs(generated_sql_dir, exist_ok=True)
            output_path = os.path.join(generated_sql_dir, f'{sanitized_name}.sql')

        # Read Excel file
        xl = pd.ExcelFile(excel_path)
        
        sql_statements = []
        tables = []
        
        # Add header comments
        sql_statements.append(f"-- Generated SQL from Excel")
        sql_statements.append(f"-- Date: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}")
        sql_statements.append(f"-- Format: {insert_format}")
        sql_statements.append(f"-- Dialect: {dialect.upper()}")
        sql_statements.append("")
        
        # Process each sheet (table)
        for sheet_name in xl.sheet_names:
            if sheet_name == '使用说明':
                continue
                
            tables.append(sheet_name)
            
            # Read sheet data
            # Explicitly use the first row as header
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=0)
            
            # Drop rows that are entirely empty (often happens at the end of user input or before separator)
            df.dropna(how='all', inplace=True)

            # Find the separator for example data by checking all cells in a row
            separator_text = "=== 示例数据 ==="
            separator_row_index = -1
            for index, row in df.iterrows():
                if separator_text in row.astype(str).values:
                    separator_row_index = index
                    break
            
            if separator_row_index != -1:
                # If separator is found, take data only from rows before the separator's row
                df = df.loc[df.index < separator_row_index]
            # If no separator, assume all rows (after header) are data meant for import.
            # This case should be rare if templates are consistently used.

            # Drop rows that are entirely empty again, in case they existed before the separator or were the only content
            df.dropna(how='all', inplace=True)

            # More robust check to remove any row that is an exact match of the column headers
            if not df.empty:
                column_header_list_str = [str(col_name).strip() for col_name in df.columns.tolist()]
                rows_to_drop = []
                for index, row_series in df.iterrows():
                    row_values_list_str = [str(item).strip() for item in row_series.tolist()]
                    if row_values_list_str == column_header_list_str:
                        rows_to_drop.append(index)
                if rows_to_drop:
                    df.drop(rows_to_drop, inplace=True)
            
            # Final check for empty DataFrame after all filtering
            if df.empty:
                continue
            
            # Generate column list
            columns = [f"`{col}`" for col in df.columns]
            column_list = ", ".join(columns)
            
            if insert_format == 'single-row':
                # Generate single-row INSERT statements
                for _, row in df.iterrows():
                    values = []
                    for col, val in zip(df.columns, row):
                        # Special handling for product_line_id to ensure it's an integer
                        if col == 'product_line_id' and isinstance(val, (int, float)) and not pd.isna(val):
                            val = int(val)
                        
                        if pd.isna(val) or (isinstance(val, str) and val.strip().upper() == 'NULL') or val == '':
                            default = FIELD_DEFAULTS.get(col, "NULL")
                            values.append(str(default))
                        elif isinstance(val, str) and (val.upper() == "NOW()" or val.upper() == "CURRENT_TIMESTAMP"):
                            values.append(val.upper())
                        elif isinstance(val, (int, float)):
                            values.append(str(val))
                        else:
                            escaped_val = str(val).replace("'", "''")
                            # Fix image paths: replace backslashes with forward slashes
                            if ('image' in col.lower() or 'url' in col.lower()) and '\\' in escaped_val:
                                escaped_val = escaped_val.replace('\\', '/')
                            values.append(f"'{escaped_val}'")
                    value_list = ", ".join(values)
                    sql_statements.append(f"INSERT INTO `{sheet_name}` ({column_list}) VALUES ({value_list});")
            else:
                # Generate multi-row INSERT statements
                rows = []
                for _, row in df.iterrows():
                    values = []
                    for col, val in zip(df.columns, row):
                        # Special handling for product_line_id to ensure it's an integer
                        if col == 'product_line_id' and isinstance(val, (int, float)) and not pd.isna(val):
                            val = int(val)
                        
                        if pd.isna(val) or (isinstance(val, str) and val.strip().upper() == 'NULL') or val == '':
                            default = FIELD_DEFAULTS.get(col, "NULL")
                            values.append(str(default))
                        elif isinstance(val, str) and (val.upper() == "NOW()" or val.upper() == "CURRENT_TIMESTAMP"):
                            values.append(val.upper())
                        elif isinstance(val, (int, float)):
                            values.append(str(val))
                        else:
                            escaped_val = str(val).replace("'", "''")
                            # Fix image paths: replace backslashes with forward slashes
                            if ('image' in col.lower() or 'url' in col.lower()) and '\\' in escaped_val:
                                escaped_val = escaped_val.replace('\\', '/')
                            values.append(f"'{escaped_val}'")
                    rows.append(f"  ({', '.join(values)})")
                
                # Split into batches
                for i in range(0, len(rows), batch_size):
                    batch = rows[i:i+batch_size]
                    values_list = ",\n".join(batch)
                    
                    sql = f"INSERT INTO `{sheet_name}` ({column_list}) VALUES\n{values_list}"
                    
                    # Add ON DUPLICATE KEY UPDATE if requested
                    if on_duplicate:
                        update_list = []
                        for col in df.columns:
                            if col.lower() != 'id':
                                update_list.append(f"`{col}` = VALUES(`{col}`)")
                        
                        if update_list:
                            sql += "\nON DUPLICATE KEY UPDATE\n  " + ",\n  ".join(update_list)
                    
                    sql += ";"
                    sql_statements.append(sql)
        
        # Update header with tables
        sql_statements[2] = f"-- Tables: {', '.join(tables)}"
        
        # Write SQL to file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(sql_statements))
        
        print(f"SQL file generated at: {output_path}")
        return output_path

def convert_sql_to_excel(sql_dir=None, output_dir=None):
    """Convert SQL files to Excel template"""
    converter = SQLExcelConverter(sql_dir, output_dir)
    converter.parse_sql_files()
    excel_path = converter.generate_excel_template()
    import_script_path = converter.generate_import_script(excel_path)
    
    if converter.warnings:
        print("\nWarnings:")
        for warning in converter.warnings:
            print(warning)
    
    return {
        'excel_path': excel_path,
        'import_script_path': import_script_path,
        'tables': list(converter.tables.keys()),
        'warnings': converter.warnings
    }

def convert_excel_to_sql(excel_path, output_dir=None, dialect='mysql', insert_format='multi-row', batch_size=1000, on_duplicate=False):
    """Convert Excel file to SQL INSERT statements"""
    converter = SQLExcelConverter(output_dir=output_dir)
    sql_path = converter.excel_to_sql(excel_path, dialect=dialect, insert_format=insert_format, 
                                     batch_size=batch_size, on_duplicate=on_duplicate)
    
    return {
        'sql_path': sql_path
    }

if __name__ == "__main__":
    # Command line interface
    import argparse
    
    parser = argparse.ArgumentParser(description='Convert between SQL and Excel')
    parser.add_argument('--mode', choices=['sql-to-excel', 'excel-to-sql'], required=True,
                        help='Conversion mode')
    parser.add_argument('--sql-dir', help='Directory containing SQL files')
    parser.add_argument('--excel-path', help='Path to Excel file')
    parser.add_argument('--output-dir', help='Output directory')
    parser.add_argument('--dialect', default='mysql', choices=['mysql', 'postgresql', 'sqlite'],
                        help='SQL dialect')
    parser.add_argument('--insert-format', default='multi-row', choices=['multi-row', 'single-row'],
                        help='INSERT statement format')
    parser.add_argument('--batch-size', type=int, default=1000,
                        help='Batch size for multi-row INSERTs')
    parser.add_argument('--on-duplicate', action='store_true',
                        help='Add ON DUPLICATE KEY UPDATE clause')
    
    args = parser.parse_args()
    
    if args.mode == 'sql-to-excel':
        result = convert_sql_to_excel(args.sql_dir, args.output_dir)
        print(f"\nConversion completed:")
        print(f"Excel template: {result['excel_path']}")
        print(f"Import script: {result['import_script_path']}")
        print(f"Tables: {', '.join(result['tables'])}")
    else:
        if not args.excel_path:
            parser.error("--excel-path is required for excel-to-sql mode")
        
        result = convert_excel_to_sql(args.excel_path, args.output_dir, args.dialect,
                                     args.insert_format, args.batch_size, args.on_duplicate)
        print(f"\nConversion completed:")
        print(f"SQL file: {result['sql_path']}") 